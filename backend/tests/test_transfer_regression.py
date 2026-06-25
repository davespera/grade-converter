from __future__ import annotations

import asyncio
import os
import zipfile
from decimal import Decimal
from io import BytesIO

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlmodel.ext.asyncio.session import AsyncSession

# Prevent import-time failures in backend.database when running tests outside Docker.
os.environ.setdefault("DATABASE_URL", "sqlite+aiosqlite:///./_placeholder_test.db")
os.environ.setdefault("FASTAPI_ACTIVEPIECES_API_KEY", "test-activepieces-key")

AUTH_HEADERS = {"x-api-key": os.environ["FASTAPI_ACTIVEPIECES_API_KEY"]}

from backend import crud, database, models
from backend.database import Base
from backend.routers import transfer


engine = create_async_engine(
    "sqlite+aiosqlite:///:memory:",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestSessionLocal = async_sessionmaker(
    engine,
    class_=AsyncSession,
    expire_on_commit=False,
    autoflush=False,
    autocommit=False,
)


async def _override_db_session():
    async with TestSessionLocal() as session:
        yield session


async def _setup_schema_and_seed(include_equivalence: bool = True) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with TestSessionLocal() as session:
        scale = models.AcademicScale(
            country_name="TESTLAND",
            scale_description="A-F",
        )
        session.add(scale)
        await session.flush()

        if include_equivalence:
            session.add(
                models.GradeEquivalence(
                    scale_id=scale.id,
                    origin_grade="A",
                    spanish_5_10=Decimal("9.00"),
                    spanish_1_4=3,
                    spanish_literal=models.SpanishLiteralEnum.SOBRESALIENTE,
                )
            )
            session.add(
                models.GradeEquivalence(
                    scale_id=scale.id,
                    origin_grade="B",
                    spanish_5_10=Decimal("8.00"),
                    spanish_1_4=3,
                    spanish_literal=models.SpanishLiteralEnum.NOTABLE,
                )
            )

        await session.commit()


async def _teardown_schema() -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
    await engine.dispose()


def test_convert_grade_returns_expected_conversion() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers=AUTH_HEADERS,
        )
        payload = response.json()

        assert response.status_code == 200
        assert "conversion" in payload
        assert isinstance(payload["conversion"], list)
        assert len(payload["conversion"]) == 1
        assert payload["conversion"][0]["origin_grade"] == "A"
        assert payload["conversion"][0]["converted_5_10"] == "9.00"
        assert payload["conversion"][0]["converted_literal"] == "SOBRESALIENTE"
        assert payload["conversion"][0]["subject"] is None
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_returns_404_for_missing_scale() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer",
            json={"scale_id": 999, "grades": [{"origin_grade": "A"}]},
            headers=AUTH_HEADERS,
        )

        assert response.status_code == 404
        assert response.json()["detail"] == "Academic scale not found"
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_returns_404_for_missing_equivalence() -> None:
    asyncio.run(_setup_schema_and_seed(include_equivalence=False))

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers=AUTH_HEADERS,
        )

        assert response.status_code == 404
        assert response.json()["detail"] == "No equivalence found for this grade"
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_returns_multiple_conversions_with_subjects() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer",
            json={
                "scale_id": 1,
                "grades": [
                    {"origin_grade": "A", "subject": "Math"},
                    {"origin_grade": "B"},
                ],
            },
            headers=AUTH_HEADERS,
        )
        payload = response.json()

        assert response.status_code == 200
        assert "conversion" in payload
        assert len(payload["conversion"]) == 2
        assert payload["conversion"][0]["origin_grade"] == "A"
        assert payload["conversion"][0]["subject"] == "Math"
        assert payload["conversion"][0]["converted_literal"] == "SOBRESALIENTE"
        assert payload["conversion"][1]["origin_grade"] == "B"
        assert payload["conversion"][1]["subject"] is None
        assert payload["conversion"][1]["converted_literal"] == "NOTABLE"
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_fetches_scale_once_for_many_grades(monkeypatch) -> None:
    # Regression: the scale lookup must happen once per request, not once per
    # grade. Three grades should still trigger a single get_scale query.
    asyncio.run(_setup_schema_and_seed())

    call_count = 0
    real_get_scale = crud.get_scale

    async def _counting_get_scale(*args, **kwargs):
        nonlocal call_count
        call_count += 1
        return await real_get_scale(*args, **kwargs)

    monkeypatch.setattr(crud, "get_scale", _counting_get_scale)

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer",
            json={
                "scale_id": 1,
                "grades": [
                    {"origin_grade": "A"},
                    {"origin_grade": "B"},
                    {"origin_grade": "A"},
                ],
            },
            headers=AUTH_HEADERS,
        )

        assert response.status_code == 200
        assert len(response.json()["conversion"]) == 3
        assert call_count == 1
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_returns_csv_file() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer?filename=grades",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers={**AUTH_HEADERS, "accept": "text/csv"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")
        assert "attachment; filename=grades.csv" in response.headers[
            "content-disposition"
        ]
        assert "origin_grade" in response.text
        assert "A" in response.text
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_defaults_to_json_when_accept_not_matched() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers={**AUTH_HEADERS, "accept": "*/*"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/json")
        assert "content-disposition" not in response.headers

        payload = response.json()
        assert payload["conversion"][0]["origin_grade"] == "A"
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_query_param_overrides_accept_header() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer?format=xlsx&filename=grades",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers={**AUTH_HEADERS, "accept": "text/csv"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment; filename=grades.xlsx" in response.headers[
            "content-disposition"
        ]

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        assert worksheet["B2"].value == "A"
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_returns_xlsx_file() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer?filename=grades",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers={
                **AUTH_HEADERS,
                "accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment; filename=grades.xlsx" in response.headers[
            "content-disposition"
        ]

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        assert worksheet["B2"].value == "A"
        assert float(worksheet["C2"].value) == 9.0
        assert worksheet["D2"].value == "SOBRESALIENTE"
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())


def test_convert_grade_returns_ods_file() -> None:
    asyncio.run(_setup_schema_and_seed())

    app = FastAPI()
    app.include_router(transfer.router)
    app.dependency_overrides[database.get_database_session] = _override_db_session

    client = None
    try:
        client = TestClient(app)
        response = client.post(
            "/transfer?filename=grades",
            json={"scale_id": 1, "grades": [{"origin_grade": "A"}]},
            headers={
                **AUTH_HEADERS,
                "accept": "application/vnd.oasis.opendocument.spreadsheet",
            },
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.oasis.opendocument.spreadsheet"
        )
        assert "attachment; filename=grades.ods" in response.headers[
            "content-disposition"
        ]

        with zipfile.ZipFile(BytesIO(response.content)) as ods_archive:
            content_xml = ods_archive.read("content.xml").decode("utf-8")
        assert "origin_grade" in content_xml
        assert ">A<" in content_xml
    finally:
        app.dependency_overrides.clear()
        if client is not None:
            client.close()
        asyncio.run(_teardown_schema())